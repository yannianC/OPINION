#!/usr/bin/env python3
"""
用 openpyxl 读取 .xlsx 文件中某一行的所有数据并打印。
用法: python read_xlsx_row.py <文件路径> <行号>
示例: python read_xlsx_row.py data.xlsx 1
"""
import sys
from openpyxl import load_workbook


def main():
    if len(sys.argv) < 3:
        print("用法: python read_xlsx_row.py <xlsx文件路径> <行号> [工作表名]")
        print("示例: python read_xlsx_row.py data.xlsx 1")
        print("示例: python read_xlsx_row.py data.xlsx 2 Sheet1")
        sys.exit(1)

    file_path = sys.argv[1]
    try:
        row_num = int(sys.argv[2])
    except ValueError:
        print("错误: 行号必须是整数")
        sys.exit(1)

    sheet_name = sys.argv[3] if len(sys.argv) > 3 else None

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
    except FileNotFoundError:
        print(f"错误: 找不到文件 '{file_path}'")
        sys.exit(1)
    except KeyError:
        print(f"错误: 工作表 '{sheet_name}' 不存在")
        sys.exit(1)

    if row_num < 1:
        print("错误: 行号从 1 开始")
        sys.exit(1)

    row_data = [cell.value for cell in ws[row_num]]
    wb.close()

    print(f"文件: {file_path}")
    print(f"工作表: {ws.title}")
    print(f"第 {row_num} 行数据 (共 {len(row_data)} 列):")
    print(row_data)

    # 也按列号打印，方便对照
    for i, val in enumerate(row_data, start=1):
        print(f"  列{i}: {val}")


if __name__ == "__main__":
    main()
